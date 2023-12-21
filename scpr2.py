from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from urllib.parse import urlparse

# Function to perform the automation
def automate_processbh(link):
    chrome_options = Options()

    # Start Chrome WebDriver with the explicit path to chromedriver.exe and options
    driver_path = r'C:\Users\fahis\OneDrive\Desktop\assesment\scraper\chromedriver.exe'
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # Open the provided link
        driver.get(link)

        # Wait for the elements to be present
        wait = WebDriverWait(driver, 10)
        h1_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1.JobDetailContent-headerTitle-rlk')))

        # Get text from the elements
        h1_text = h1_element.text
        return h1_text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None

    finally:
        # Close the browser window
        driver.quit()
     
def automate_processuw(link):
    chrome_options = Options()

    # Start Chrome WebDriver with the explicit path to chromedriver.exe and options
    driver_path = r'C:\Users\fahis\OneDrive\Desktop\assesment\scraper\chromedriver.exe'
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # Open the provided link
        driver.get(link)

        # Wait for the elements to be present
        wait = WebDriverWait(driver, 10)
        h1_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h4.m-0')))


        # Get text from the elements
        h1_text = h1_element.text

        return h1_text

    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None

    finally:
        # Close the browser window
        driver.quit()
     
def automate_process(link):
    # Chrome options
    chrome_options = Options()

    # Start Chrome WebDriver with the explicit path to chromedriver.exe and options
    driver_path = r'C:\Users\fahis\OneDrive\Desktop\assesment\scraper\chromedriver.exe'
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # Open the provided link
        driver.get(link)

        # Wait for the elements to be present
        wait = WebDriverWait(driver, 10)
        h1_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1.PageProjectViewLogout-projectInfo-title')))
        p_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'p.PageProjectViewLogout-detail-projectId')))
        # bid_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input#form-signup-bottom-bid')))

        # Get text from the elements
        h1_text = h1_element.text
        p_text = p_element.text
        # bid = bid_input.get_attribute("placeholder")
        p_text = p_text.split('#')[1].strip()
        return h1_text, p_text

    except Exception as e:
        print(f"An error occurred: {e}")
        return "error", 0

    finally:
        # Close the browser window
        driver.quit()

# Read links from links.txt
links = []
with open('links.txt', 'r') as file:
    links = file.readlines()

# Load the existing template Excel file
wb = openpyxl.load_workbook(r'C:\Users\fahis\OneDrive\Desktop\assesment\scraper\project.xlsx')
sheet = wb.active

# Process each link and insert data into respective columns
row = 32  # Starting from row 2, assuming header in row 1
for link in links:
    parsed_url = urlparse(link)
    domain_parts = parsed_url.netloc.split('.')
    main_domain = domain_parts[-2] if len(domain_parts) > 2 else domain_parts[0]

    if(main_domain=='freelancer'):
        link = link.strip()
      # Remove leading/trailing whitespace, newlines, etc.
        h1_text, p_text = automate_process(link)
        sheet.cell(row=row, column=4).value = h1_text
        sheet.cell(row=row, column=3).value = 'freelancer'
        sheet.cell(row=row, column=2).value = '2.11.2023'
        sheet.cell(row=row, column=6).value = int(p_text)
        sheet.cell(row=row, column=1).value = row-9
        sheet.cell(row=row, column=5).value = link
        # sheet.cell(row=row, column=7).value = int(bid) if bid else ""
        row += 1
        print(row, "finished fl")
        wb.save('project.xlsx')
    elif(main_domain=='behance'):
            link = link.strip()
      # Remove leading/trailing whitespace, newlines, etc.
            h1_text = automate_processbh(link)
            sheet.cell(row=row, column=4).value = h1_text
            sheet.cell(row=row, column=5).value = link
            sheet.cell(row=row, column=2).value = '2.11.2023'
            sheet.cell(row=row, column=3).value = 'behance'
            sheet.cell(row=row, column=1).value = row-9
            row += 1
            print(row, "finished bh")
            wb.save('project.xlsx')
    elif(main_domain=='upwork'):
            link = link.strip()
        # Remove leading/trailing whitespace, newlines, etc.
            h1_text = automate_processuw(link)
            sheet.cell(row=row, column=4).value = h1_text
            sheet.cell(row=row, column=5).value = link
            sheet.cell(row=row, column=2).value = '2.11.2023'
            sheet.cell(row=row, column=3).value = 'upwork'
            sheet.cell(row=row, column=1).value = row-9
            row += 1
            print(row, "finished uw")
            wb.save('project.xlsx')



# Save the modified Excel file
print("EXCEL SAVED")
wb.save('project.xlsx')
